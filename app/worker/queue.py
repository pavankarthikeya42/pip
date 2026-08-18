"""Excel import and job creation.

The KARI PIP export has a report header before the real table header.
This importer detects the header row instead of assuming row 1.
"""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from ..database.database import Database


EXPECTED_HEADERS = {
    "Brand Name": "brand_name",
    "Generic Name": "generic_name",
    "Sponsor": "sponsor",
    "PIP Number": "pip_number",
    "Decision Type": "decision_type",
    "Decision Date": "decision_date",
    "Status": "status",
    "Therapeutic Areas": "therapeutic_areas",
    "Decision Number": "decision_number",
    "Condition / Indication": "condition_indication",
    "Modifications Scope": "modifications_scope",
    "Modification Decision #": "modification_decision_number",
    "Modifications Dates": "modifications_dates",
    "Discontinue Date": "discontinue_date",
    "Discontinue Reason": "discontinue_reason",
    "Routes of Administration": "routes_of_administration",
    "Pharmaceutical Forms": "pharmaceutical_forms",
    "First Published": "first_published",
    "Last Updated": "last_updated",
}


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def detect_header_row(ws, scan_rows: int = 20) -> tuple[int, dict[str, int]]:
    """Find the row containing the PIP export's actual column headers."""
    expected = {h.casefold(): key for h, key in EXPECTED_HEADERS.items()}
    best_row = None
    best_matches: dict[str, int] = {}

    for row_idx in range(1, min(ws.max_row, scan_rows) + 1):
        matches: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            value = _stringify(ws.cell(row_idx, col_idx).value)
            key = expected.get(value.casefold())
            if key:
                matches[key] = col_idx
        if len(matches) > len(best_matches):
            best_row, best_matches = row_idx, matches

    required = {"brand_name", "generic_name", "pip_number"}
    if best_row is None or not required.issubset(best_matches):
        raise ValueError(
            "Could not locate the PIP metadata header row. "
            f"Expected at least: {sorted(required)}"
        )

    return best_row, best_matches


def load_jobs_from_excel(
    db: Database,
    path: str | Path,
    owner: str,
    limit: int | None = None,
    medicine: str | None = None,
) -> int:
    """Import the supplied PIP Excel without splitting it.

    The importer preserves all 19 metadata fields in metadata_json.  The
    Brand Name is used as medicine_name because this export does not contain
    a separate Medicine Name column.
    """
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    ws = wb.active
    header_row, header_cols = detect_header_row(ws)

    rows: list[dict[str, Any]] = []
    medicine_filter = medicine.casefold().strip() if medicine else None

    for row_idx in range(header_row + 1, ws.max_row + 1):
        metadata: dict[str, str] = {}
        for field, col_idx in header_cols.items():
            metadata[field] = _stringify(ws.cell(row_idx, col_idx).value)

        medicine_name = metadata.get("brand_name", "").strip()
        if not medicine_name:
            continue
        if medicine_filter and medicine_name.casefold() != medicine_filter:
            continue

        metadata["medicine_name"] = medicine_name
        metadata["source_file"] = path.name
        metadata["source_row"] = str(row_idx)
        metadata["owner"] = owner
        rows.append(metadata)

        if limit is not None and len(rows) >= limit:
            break

    if not rows:
        raise ValueError(
            f"No medicine records found in {path.name}"
            + (f" for medicine '{medicine}'" if medicine else "")
        )

    db.upsert_medicines(rows)
    db.create_jobs(owner=owner)
    return len(rows)


def stream_jobs_from_excel(
    path: str | Path,
    limit: int | None = None,
    medicine: str | None = None,
) -> list[dict[str, Any]]:
    """Stream PIP job records directly from Excel without using a database."""
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    ws = wb.active
    header_row, header_cols = detect_header_row(ws)

    rows: list[dict[str, Any]] = []
    medicine_filter = medicine.casefold().strip() if medicine else None

    for row_idx in range(header_row + 1, ws.max_row + 1):
        metadata: dict[str, str] = {}
        for field, col_idx in header_cols.items():
            metadata[field] = _stringify(ws.cell(row_idx, col_idx).value)

        medicine_name = metadata.get("brand_name", "").strip()
        if not medicine_name:
            continue
        if medicine_filter and medicine_name.casefold() != medicine_filter:
            continue

        metadata["medicine_name"] = medicine_name
        metadata["source_file"] = path.name
        metadata["source_row"] = str(row_idx)
        rows.append(metadata)

        if limit is not None and len(rows) >= limit:
            break

    if not rows:
        raise ValueError(
            f"No medicine records found in {path.name}"
            + (f" for medicine '{medicine}'" if medicine else "")
        )

    return rows

